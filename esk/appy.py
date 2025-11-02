import os
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

ENV_BATCH_WORKBOOK = "ESKER_INVOICE_BATCH_WORKBOOK"


def resolve_workbook() -> Path:
    """Locate the source workbook that contains the per-worksheet invoice lists."""
    candidates: List[Path] = []

    env_candidate = os.environ.get(ENV_BATCH_WORKBOOK)
    if env_candidate:
        candidates.append(Path(env_candidate).expanduser())

    script_dir = Path(__file__).resolve().parent
    candidates.append(script_dir / "invoice_n.xlsx")
    candidates.append(Path.home() / "Downloads" / "invoice_n.xlsx")

    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate

    raise FileNotFoundError(
        f"Unable to locate 'invoice_n.xlsx'. Set {ENV_BATCH_WORKBOOK} to the workbook path."
    )


def resolve_username() -> str:
    """Get the Windows username once so child runs skip the Tkinter prompt."""
    username = os.environ.get("ESKER_USERNAME", "").strip()
    if username:
        return username

    while True:
        try:
            username = input("Enter Esker Windows username (for Downloads folder access): ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nUsername is required. Aborting.")
            sys.exit(1)

        if username:
            return username

        print("Username cannot be empty. Please try again.")


def load_sheet_names(workbook_path: Path) -> List[str]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def prepare_invoice_dataframe(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(workbook_path, sheet_name=sheet_name, engine="openpyxl")

    if "invoice" not in df.columns:
        raise KeyError("required column 'invoice' is missing")

    df = df.copy()
    df = df.dropna(subset=["invoice"])
    df["invoice"] = df["invoice"].astype(str).str.strip()
    df = df[df["invoice"] != ""]
    df = df.drop_duplicates(subset=["invoice"])

    return df[["invoice"]]


def sanitize_for_fs(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_")
    return cleaned or "sheet"


def run_sheet(
    app_path: Path,
    python_executable: Path,
    base_env: Dict[str, str],
    sheet_name: str,
    df_invoice: pd.DataFrame,
) -> None:
    safe_sheet = sanitize_for_fs(sheet_name)

    with tempfile.TemporaryDirectory(prefix=f"invoice_{safe_sheet}_") as tmp_dir:
        invoice_file = Path(tmp_dir) / "invoice.xlsx"

        with pd.ExcelWriter(invoice_file, engine="openpyxl") as writer:
            df_invoice.to_excel(writer, sheet_name="invoice", index=False)

        child_env = base_env.copy()
        child_env["ESKER_INVOICE_WORKBOOK"] = str(invoice_file)
        child_env["ESKER_INVOICE_SHEET"] = "invoice"

        subprocess.run(
            [str(python_executable), str(app_path)],
            check=True,
            cwd=str(app_path.parent),
            env=child_env,
        )


def main() -> int:
    try:
        workbook_path = resolve_workbook()
    except FileNotFoundError as exc:
        print(exc)
        return 1

    app_path = Path(__file__).resolve().with_name("app.py")
    if not app_path.exists():
        print(f"Automation script not found at {app_path}.")
        return 1

    try:
        sheet_names = load_sheet_names(workbook_path)
    except Exception as exc:
        print(f"Unable to read workbook '{workbook_path}': {exc}")
        return 1

    if not sheet_names:
        print(f"No worksheets found in '{workbook_path}'. Nothing to do.")
        return 0

    python_executable = Path(sys.executable).resolve()
    username = resolve_username()

    base_env = os.environ.copy()
    base_env["ESKER_USERNAME"] = username

    completed: List[str] = []
    skipped: List[Tuple[str, str]] = []
    failed: List[Tuple[str, str]] = []

    print(f"Found {len(sheet_names)} worksheet(s) in '{workbook_path}'.")

    for sheet_name in sheet_names:
        try:
            df_invoice = prepare_invoice_dataframe(workbook_path, sheet_name)
        except KeyError as exc:
            message = str(exc)
            skipped.append((sheet_name, message))
            print(f"Skipping sheet '{sheet_name}': {message}")
            continue
        except Exception as exc:
            failed.append((sheet_name, f"read error: {exc}"))
            print(f"Failed to read sheet '{sheet_name}': {exc}")
            continue

        if df_invoice.empty:
            skipped.append((sheet_name, "no invoice rows after cleaning"))
            print(f"Skipping sheet '{sheet_name}': no invoice rows after cleaning.")
            continue

        print(f"Launching Selenium workflow for sheet '{sheet_name}' ({len(df_invoice)} invoices).")

        try:
            run_sheet(app_path, python_executable, base_env, sheet_name, df_invoice)
        except subprocess.CalledProcessError as exc:
            failed.append((sheet_name, f"return code {exc.returncode}"))
            print(f"Sheet '{sheet_name}' failed with return code {exc.returncode}.")
        except Exception as exc:
            failed.append((sheet_name, str(exc)))
            print(f"Sheet '{sheet_name}' failure: {exc}")
        else:
            completed.append(sheet_name)

    print("\nSummary")
    if completed:
        print(f"  Completed ({len(completed)}): {', '.join(completed)}")
    else:
        print("  Completed (0)")

    if skipped:
        details = "; ".join(f"{name} -> {reason}" for name, reason in skipped)
        print(f"  Skipped ({len(skipped)}): {details}")
    else:
        print("  Skipped (0)")

    if failed:
        details = "; ".join(f"{name} -> {reason}" for name, reason in failed)
        print(f"  Failed ({len(failed)}): {details}")
        return 1

    print("  Failed (0)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
